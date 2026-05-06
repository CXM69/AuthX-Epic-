from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import re
from typing import BinaryIO, Iterable

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ACCOUNT_COLUMN_CANDIDATES = [
    "account name",
    "account",
    "organization",
    "organization name",
    "org name",
    "health system",
    "health system name",
    "system",
    "system name",
    "customer",
    "customer name",
    "name",
]

PARENT_COLUMN_CANDIDATES = [
    "parent health system",
    "parent system",
    "parent organization",
    "parent",
    "health system parent",
    "ultimate parent",
]

EHR_COLUMN_CANDIDATES = [
    "ehr",
    "emr",
    "vendor",
    "ehr vendor",
    "emr vendor",
    "platform",
    "source ehr",
]

STATUS_COLUMN_CANDIDATES = [
    "status",
    "customer status",
    "relationship",
    "stage",
    "notes",
]

STATE_COLUMN_CANDIDATES = ["state", "st", "region"]

EXCLUDE_TERMS = {"test", "demo", "sample", "training", "unknown", "n/a", "na", "none"}

LEGAL_SUFFIX_PATTERN = re.compile(
    r"\b(?:inc|incorporated|llc|l\.l\.c|ltd|limited|corp|corporation|co|company|plc)\b",
    re.IGNORECASE,
)

ORG_WORD_PATTERN = re.compile(
    r"\b(?:the|healthcare|health care|health system|health systems|hospital system|hospitals|hospital|medical center|medical ctr|clinic|clinics|network)\b",
    re.IGNORECASE,
)

ORGANIZATION_SIGNAL_PATTERN = re.compile(
    r"\b(?:health|hospital|medical|clinic|system|systems|care|center|centre|regional|memorial|university|children|county|network|partners|group|foundation)\b",
    re.IGNORECASE,
)

NEW_EPIC_PATTERN = re.compile(
    r"\b(?:new epic|epic implementation|implementing epic|epic go live|epic go-live|go live|go-live|installing epic|epic migration|migrating to epic|epic rollout)\b",
    re.IGNORECASE,
)

NON_EPIC_EHR_PATTERN = re.compile(
    r"\b(?:cerner|oracle health|meditech|allscripts|altera|athena|athenahealth|eclinicalworks|ecw|nextgen|greenway|veradigm|scripts)\b",
    re.IGNORECASE,
)

MIGRATION_PATTERN = re.compile(
    r"\b(?:migration|migrate|replacement|replace|legacy|modernization|consolidation|rationalization|transition|upgrade)\b",
    re.IGNORECASE,
)

IMPRIVATA_PATTERN = re.compile(
    r"\b(?:imprivata|onesign|one sign|confirm id|fairwarning|groundcontrol)\b",
    re.IGNORECASE,
)

LARGE_HEALTH_SYSTEM_PATTERN = re.compile(
    r"\b(?:idn|integrated delivery|enterprise|multi hospital|multi-hospital|large health|regional health|academic medical|university health|health system|hospital network)\b",
    re.IGNORECASE,
)

LARGE_PHYSICIAN_GROUP_PATTERN = re.compile(
    r"\b(?:physician group|medical group|provider group|multi specialty|multi-specialty|ambulatory group|clinic network|providers|physicians)\b",
    re.IGNORECASE,
)

FRONT_DOOR_PATTERN = re.compile(
    r"\b(?:front door|front-door|workstation|shared workstation|badge|tap|tap in|tap out|sso|single sign|mfa|multi factor|multifactor|biometric|biometrics|passkey|passkeys|vdi|thin client|thin-client|step up|step-up|reauth|re-auth|authentication|access workflow)\b",
    re.IGNORECASE,
)

SECURITY_PATTERN = re.compile(
    r"\b(?:ciso|security|mfa|multi factor|multifactor|compliance|hipaa|audit|risk|zero trust|ransomware|cyber|breach|access control)\b",
    re.IGNORECASE,
)

BUYER_ROLE_PATTERN = re.compile(
    r"\b(?:cio|ciso|iam|identity|epic director|epic program|clinical applications|cmio|cnio|pharmacy informatics|end user computing|desktop|euc)\b",
    re.IGNORECASE,
)

VDI_PATTERN = re.compile(r"\b(?:vdi|thin client|thin-client|shared workstation|workstation|desktop|euc|end user computing)\b", re.IGNORECASE)
CLINICAL_FRICTION_PATTERN = re.compile(r"\b(?:clinician friction|workflow friction|physician friction|nurse workflow|clinical workflow|provider experience|clinician experience)\b", re.IGNORECASE)
EPCS_PATTERN = re.compile(r"\b(?:epcs|pharmacy|pharmacist|break glass|break-glass|controlled substance)\b", re.IGNORECASE)
MODERNIZATION_PATTERN = re.compile(r"\b(?:enterprise modernization|modernization|standardization|consolidation|digital transformation|access modernization)\b", re.IGNORECASE)

AUTHX_GUARDRAIL = (
    "Position AuthX as the authentication and access layer around Epic: workstation access, SSO, MFA, "
    "badge tap, biometrics, passkeys, VDI/thin-client access, and step-up re-authentication. "
    "Do not position AuthX as replacing Epic role assignment, identity governance, provisioning, "
    "deprovisioning, or internal Epic authorization."
)


@dataclass(frozen=True)
class SourceConfig:
    label: str
    uploaded_file: BinaryIO
    source_type: str


def infer_source_type(label: str) -> str:
    normalized = normalize_column_name(label)

    if "imprivata" in normalized:
        return "Imprivata Customer List"
    if "health systems by ehr" in normalized or "ehr" in normalized or "emr" in normalized:
        return "Health Systems by EHR"
    if "epic organization" in normalized or "epic org" in normalized or "epic customer" in normalized:
        return "EPIC Organization List"
    if "new epic" in normalized or "epic implementation" in normalized or "epic migration" in normalized:
        return "New Epic Systems List"
    if "epic" in normalized and "list" in normalized:
        return "EPIC Organization List"
    return "Uploaded Workbook"


def source_configs_from_uploads(uploaded_files: Iterable[BinaryIO]) -> list[SourceConfig]:
    sources: list[SourceConfig] = []
    for index, uploaded_file in enumerate(uploaded_files, start=1):
        label = clean_text(getattr(uploaded_file, "name", "")) or f"Workbook {index}"
        sources.append(
            SourceConfig(
                label=label,
                uploaded_file=uploaded_file,
                source_type=infer_source_type(label),
            )
        )
    return sources


def normalize_column_name(column: object) -> str:
    text = str(column).strip().lower()
    text = re.sub(r"[\s_\-/]+", " ", text)
    return re.sub(r"[^a-z0-9 ]+", "", text).strip()


def normalize_account_name(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip().lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = LEGAL_SUFFIX_PATTERN.sub(" ", text)
    text = ORG_WORD_PATTERN.sub(" ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    text = re.sub(r"\s+", " ", str(value).strip())
    return "" if text.lower() in {"nan", "none", "null"} else text


def read_all_tabs(source: SourceConfig) -> pd.DataFrame:
    if hasattr(source.uploaded_file, "seek"):
        source.uploaded_file.seek(0)

    workbook = pd.ExcelFile(source.uploaded_file, engine="openpyxl")
    frames: list[pd.DataFrame] = []

    for sheet_name in workbook.sheet_names:
        frame = read_sheet_with_header_detection(workbook, sheet_name).dropna(how="all")
        if frame.empty:
            continue
        frame.columns = [str(column).strip() for column in frame.columns]
        frame["Source File"] = source.label
        frame["Source Tab"] = sheet_name
        frame["Source Type"] = source.source_type
        frames.append(frame)

    return pd.concat(frames, ignore_index=True, sort=False) if frames else pd.DataFrame()


def read_sheet_with_header_detection(workbook: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    raw = pd.read_excel(workbook, sheet_name=sheet_name, header=None, dtype=object)
    raw = raw.dropna(axis=0, how="all").dropna(axis=1, how="all")
    if raw.empty:
        return pd.DataFrame()

    header_row = detect_header_row(raw)
    if header_row is None:
        frame = raw.copy()
        frame.columns = [f"Column {index + 1}" for index in range(len(frame.columns))]
        return frame

    frame = raw.iloc[header_row + 1 :].copy()
    frame.columns = make_unique_columns(raw.iloc[header_row].tolist())
    return frame.dropna(axis=0, how="all")


def detect_header_row(raw: pd.DataFrame, max_rows: int = 15) -> int | None:
    candidates = (
        ACCOUNT_COLUMN_CANDIDATES
        + PARENT_COLUMN_CANDIDATES
        + EHR_COLUMN_CANDIDATES
        + STATUS_COLUMN_CANDIDATES
        + STATE_COLUMN_CANDIDATES
    )
    normalized_candidates = {normalize_column_name(candidate) for candidate in candidates}
    best_row: int | None = None
    best_score = 0

    for row_index in range(min(max_rows, len(raw))):
        values = [normalize_column_name(value) for value in raw.iloc[row_index].tolist()]
        values = [value for value in values if value and not value.startswith("unnamed")]
        score = 0
        for value in values:
            if value in normalized_candidates:
                score += 5
            elif any(candidate in value for candidate in normalized_candidates if len(candidate) > 2):
                score += 2
        if score > best_score:
            best_score = score
            best_row = row_index

    return best_row if best_score else None


def make_unique_columns(values: list[object]) -> list[str]:
    columns: list[str] = []
    counts: dict[str, int] = {}

    for index, value in enumerate(values):
        column = clean_text(value) or f"Column {index + 1}"
        counts[column] = counts.get(column, 0) + 1
        if counts[column] > 1:
            column = f"{column} {counts[column]}"
        columns.append(column)

    return columns


def find_column(frame: pd.DataFrame, candidates: list[str]) -> str | None:
    normalized_candidates = [normalize_column_name(candidate) for candidate in candidates]
    normalized_lookup = {
        normalize_column_name(column): column
        for column in frame.columns
        if not normalize_column_name(column).startswith("unnamed")
    }

    for candidate in normalized_candidates:
        if candidate in normalized_lookup:
            return normalized_lookup[candidate]

    for column in frame.columns:
        normalized = normalize_column_name(column)
        tokens = set(normalized.split())
        for candidate in normalized_candidates:
            if len(candidate) <= 2 and candidate in tokens:
                return column
            if len(candidate) > 2 and candidate in normalized:
                return column

    return None


def infer_account_column(frame: pd.DataFrame) -> str | None:
    best_column: str | None = None
    best_score = 0.0

    for column in frame.columns:
        if normalize_column_name(column) in {"source file", "source tab", "source type"}:
            continue
        values = frame[column].dropna().astype(str).str.strip()
        values = values[values.ne("") & ~values.str.lower().isin({"nan", "none", "null", "n/a", "na"})].head(300)
        if values.empty:
            continue
        useful_values = values[values.str.len().between(3, 120) & ~values.str.fullmatch(r"[-+]?\d+(\.\d+)?", na=False)]
        if useful_values.empty:
            continue
        signal_hits = useful_values.str.contains(ORGANIZATION_SIGNAL_PATTERN, na=False).sum()
        score = (len(useful_values) * 2) + (signal_hits * 8) + (useful_values.nunique() * 2)
        if score > best_score:
            best_score = score
            best_column = column

    return best_column


def row_text(frame: pd.DataFrame) -> pd.Series:
    return frame.apply(
        lambda row: " ".join(clean_text(value) for value in row.tolist() if clean_text(value)),
        axis=1,
    )


def prepare_source(frame: pd.DataFrame, source_type: str) -> pd.DataFrame:
    output_columns = [
        "Account Name",
        "Normalized Account Name",
        "Parent Health System if available",
        "EHR",
        "Status",
        "State",
        "Source File",
        "Source Tab",
        "Source Type",
        "Row Text",
    ]
    if frame.empty:
        return pd.DataFrame(columns=output_columns)

    account_column = find_column(frame, ACCOUNT_COLUMN_CANDIDATES) or infer_account_column(frame)
    if account_column is None:
        raise ValueError(f"Could not find a usable account/name column in {source_type}.")

    parent_column = find_column(frame, PARENT_COLUMN_CANDIDATES)
    ehr_column = find_column(frame, EHR_COLUMN_CANDIDATES)
    status_column = find_column(frame, STATUS_COLUMN_CANDIDATES)
    state_column = find_column(frame, STATE_COLUMN_CANDIDATES)

    prepared = pd.DataFrame()
    prepared["Account Name"] = frame[account_column].map(clean_text)
    prepared["Normalized Account Name"] = prepared["Account Name"].map(normalize_account_name)
    if parent_column and parent_column != account_column:
        prepared["Parent Health System if available"] = frame[parent_column].map(clean_text)
    else:
        prepared["Parent Health System if available"] = ""
    prepared["EHR"] = frame[ehr_column].map(clean_text) if ehr_column else ""
    prepared["Status"] = frame[status_column].map(clean_text) if status_column else ""
    prepared["State"] = frame[state_column].map(clean_text) if state_column else ""
    prepared["Source File"] = frame.get("Source File", "")
    prepared["Source Tab"] = frame.get("Source Tab", "")
    prepared["Source Type"] = source_type
    prepared["Row Text"] = row_text(frame)

    prepared = prepared[prepared["Normalized Account Name"].ne("")]
    prepared = prepared[
        ~prepared["Normalized Account Name"].isin(EXCLUDE_TERMS)
        & prepared["Account Name"].str.len().gt(1)
    ]
    return prepared[output_columns].reset_index(drop=True)


def collapse_values(series: pd.Series, limit: int = 8) -> str:
    values = [
        clean_text(value)
        for value in series.dropna().tolist()
        if clean_text(value) and clean_text(value).lower() not in {"nan", "none"}
    ]
    return "; ".join(list(dict.fromkeys(values))[:limit])


def has_pattern(text: str, pattern: re.Pattern[str]) -> bool:
    return bool(pattern.search(text or ""))


def score_account(row: pd.Series) -> pd.Series:
    text = row["Source Text"]

    new_epic = has_pattern(text, NEW_EPIC_PATTERN)
    epic_customer = bool(row["Existing Epic Customer"]) or new_epic
    non_epic_migration = bool(row["Non-Epic Migration Opportunity"])
    imprivata = bool(row["Imprivata Customer"])
    large_health_system = bool(row["Large Health System"])
    large_physician_group = bool(row["Large Physician Group"])
    front_door = bool(row["Front-Door Workflow Fit Signal"])
    security = bool(row["Security / Compliance Signal"])
    buyer_role = bool(row["Buyer Role Signal"])

    epic_status_points = 25 if epic_customer else 10 if non_epic_migration else 0
    timing_points = 20 if new_epic else 10 if has_pattern(text, MIGRATION_PATTERN) else 0
    enterprise_points = 15 if large_health_system else 12 if large_physician_group else min(int(row["Source Count"]) * 4, 8)
    competitive_points = 15 if imprivata else 10 if has_pattern(text, re.compile(r"\b(?:duo|okta|ping|mfa|sso|single sign)\b", re.IGNORECASE)) else 0
    front_door_points = 10 if front_door else 5 if epic_customer else 0
    security_points = 10 if security else 0
    buyer_points = 5 if buyer_role else 3 if epic_customer or large_health_system else 0

    score = (
        epic_status_points
        + timing_points
        + enterprise_points
        + competitive_points
        + front_door_points
        + security_points
        + buyer_points
    )
    if bool(row["Exclude Flag"]):
        score = 0

    return pd.Series(
        {
            "Epic Status Points": epic_status_points,
            "Timing Trigger Points": timing_points,
            "Enterprise Value Points": enterprise_points,
            "Imprivata / Competitive Signal Points": competitive_points,
            "Front-Door Workflow Fit Points": front_door_points,
            "Security / Compliance Urgency Points": security_points,
            "Buyer Role Fit Points": buyer_points,
            "AuthX Score": max(0, min(100, int(score))),
        }
    )


def assign_tier(score: int) -> str:
    if score >= 85:
        return "Tier 1"
    if score >= 70:
        return "Tier 2"
    if score >= 50:
        return "Tier 3"
    if score >= 30:
        return "Hold"
    return "Exclude"


def epic_status(row: pd.Series) -> str:
    if row["New Epic System"]:
        return "New Epic Implementation"
    if row["Existing Epic Customer"]:
        return "Existing Epic Customer"
    if row["Non-Epic Migration Opportunity"]:
        return "Non-Epic Migration Opportunity"
    return "Non-Epic / Unknown"


def imprivata_signal(row: pd.Series) -> str:
    return "Imprivata signal found" if row["Imprivata Customer"] else "No Imprivata signal found"


def why_account_matters(row: pd.Series) -> str:
    reasons: list[str] = []
    if row["New Epic System"]:
        reasons.append("new Epic timing trigger")
    elif row["Existing Epic Customer"]:
        reasons.append("existing Epic footprint")
    if row["Imprivata Customer"]:
        reasons.append("Imprivata authentication signal")
    if row["Large Health System"]:
        reasons.append("large health system value")
    if row["Large Physician Group"]:
        reasons.append("large physician group value")
    if row["Non-Epic Migration Opportunity"]:
        reasons.append("possible non-Epic migration opportunity")
    if row["Front-Door Workflow Fit Signal"]:
        reasons.append("front-door workflow fit")
    if row["Security / Compliance Signal"]:
        reasons.append("security or compliance urgency")
    return "Prioritize because of " + ", ".join(reasons) + "." if reasons else "Limited explicit signal; keep for research."


def recommend_role(row: pd.Series) -> pd.Series:
    text = row["Source Text"]

    if row["New Epic System"]:
        role = "Epic Program Director"
        why = "This is a new Epic system, so authentication workflow decisions are likely still being made before go-live."
        secondary = "CISO; IAM Director; CMIO; CNIO; Director of End User Computing"
        pitch = "Let's validate your Epic front-door authentication model before go-live: workstation access, SSO, badge tap, MFA, and step-up re-authentication."
        action = "Send Epic Front-Door Authentication Audit outreach."
    elif row["Existing Epic Customer"]:
        role = "Epic Director or VP Clinical Applications"
        why = "This account already has Epic, so the clinical applications owner is likely closest to Epic access friction."
        secondary = "IAM Director; CISO; CMIO; CNIO; Director of End User Computing"
        pitch = "AuthX can strengthen the authentication layer around Epic without changing Epic role assignment or internal authorization."
        action = "Request a short Epic access workflow review."
    elif row["Security / Compliance Signal"]:
        role = "CISO"
        why = "Security, MFA, compliance, or access-control urgency appears in the account signals."
        secondary = "IAM Director; CIO; Epic Director; VP Clinical Applications"
        pitch = "AuthX helps enforce MFA, step-up re-authentication, badge tap, biometrics, and secure access around Epic workflows."
        action = "Send security-led front-door authentication risk outreach."
    elif row["Imprivata Customer"]:
        role = "IAM Director"
        why = "Imprivata signal suggests identity and clinical authentication workflows are already owned and funded."
        secondary = "CISO; Epic Director; VP Clinical Applications; Director of End User Computing"
        pitch = "AuthX can complement clinical SSO and access workflows around Epic with MFA, badge tap, biometrics, passkeys, and step-up controls."
        action = "Ask who owns clinical IAM and Imprivata workflow strategy."
    elif has_pattern(text, VDI_PATTERN):
        role = "Director of End User Computing"
        why = "VDI, thin-client, shared-workstation, or desktop signals point to front-door access ownership."
        secondary = "IAM Director; CISO; Epic Director; Desktop Engineering Leader"
        pitch = "AuthX helps secure workstation, VDI, thin-client, badge tap, and shared clinical access into Epic."
        action = "Send shared-workstation and VDI access workflow outreach."
    elif has_pattern(text, CLINICAL_FRICTION_PATTERN):
        role = "CMIO or CNIO"
        why = "Clinical workflow friction signals point to clinical leadership impact and adoption risk."
        secondary = "Epic Director; VP Clinical Applications; IAM Director; CISO"
        pitch = "AuthX reduces clinical authentication friction around Epic while preserving Epic authorization and role controls."
        action = "Lead with clinician workflow friction and request a clinical access conversation."
    elif has_pattern(text, EPCS_PATTERN):
        role = "Pharmacy Informatics Director or CMIO"
        why = "EPCS, pharmacy, or break-glass signals point to high-friction step-up authentication workflows."
        secondary = "CISO; IAM Director; Epic Security Manager; VP Clinical Applications"
        pitch = "AuthX can support step-up re-authentication for sensitive Epic workflows like EPCS, pharmacy, and break-glass access."
        action = "Send EPCS and step-up authentication workflow outreach."
    else:
        role = "CIO"
        why = "Enterprise modernization or limited role-specific data means IT leadership is the best first routing point."
        secondary = "CISO; IAM Director; Epic Director; VP Clinical Applications"
        pitch = "AuthX provides the authentication and access layer around Epic for SSO, MFA, badge tap, biometrics, passkeys, VDI, and step-up re-authentication."
        action = "Send executive front-door authentication discovery outreach."

    return pd.Series(
        {
            "Best Role to Pursue First": role,
            "Why This Role": why,
            "Secondary Roles": secondary,
            "Best Pitch": pitch,
            "Next Sales Action": action,
        }
    )


def confidence(row: pd.Series) -> str:
    if row["AuthX Score"] >= 85 and row["Source Count"] >= 2:
        return "High"
    if row["AuthX Score"] >= 70 or row["New Epic System"] or (row["Existing Epic Customer"] and row["Imprivata Customer"]):
        return "Medium"
    return "Low"


def missing_data(row: pd.Series) -> str:
    missing: list[str] = []
    if not row["Parent Health System if available"]:
        missing.append("parent health system")
    if not row["EHR"]:
        missing.append("EHR")
    if not row["New Epic System"]:
        missing.append("Epic timing")
    if not row["Imprivata Customer"]:
        missing.append("Imprivata signal")
    if not row["Buyer Role Signal"]:
        missing.append("named buyer role")
    return "; ".join(missing) if missing else "None"


def build_ranked_accounts_from_sources(sources: Iterable[SourceConfig]) -> tuple[pd.DataFrame, dict[str, int]]:
    source_list = list(sources)
    if not source_list:
        raise ValueError("Upload at least one Excel workbook before running scoring.")

    prepared: list[pd.DataFrame] = []
    skipped_sources: list[str] = []

    for source in source_list:
        try:
            source_frame = read_all_tabs(source)
            prepared_frame = prepare_source(source_frame, source.source_type)
        except ValueError as exc:
            skipped_sources.append(f"{source.label}: {exc}")
            continue

        if prepared_frame.empty:
            skipped_sources.append(f"{source.label}: no usable account rows found")
            continue
        prepared.append(prepared_frame)

    if not prepared:
        details = f" Skipped sources: {' | '.join(skipped_sources[:3])}" if skipped_sources else ""
        raise ValueError(f"No usable account rows were found in the uploaded workbooks.{details}")

    combined = pd.concat(prepared, ignore_index=True, sort=False)

    if combined.empty:
        raise ValueError("No usable account rows were found in the uploaded workbooks.")

    combined["Data Text"] = (
        combined["Account Name"].fillna("").astype(str)
        + " "
        + combined["Parent Health System if available"].fillna("").astype(str)
        + " "
        + combined["EHR"].fillna("").astype(str)
        + " "
        + combined["Status"].fillna("").astype(str)
        + " "
        + combined["Row Text"].fillna("").astype(str)
    ).str.lower()
    combined["Search Text"] = (
        combined["Data Text"].fillna("").astype(str)
        + " "
        + combined["Source File"].fillna("").astype(str)
        + " "
        + combined["Source Tab"].fillna("").astype(str)
        + " "
        + combined["Source Type"].fillna("").astype(str)
    ).str.lower()

    grouped = combined.groupby("Normalized Account Name", dropna=False)
    ranked = grouped.agg(
        Account_Name=("Account Name", "first"),
        Parent_Health_System_if_available=("Parent Health System if available", collapse_values),
        EHR=("EHR", collapse_values),
        Status=("Status", collapse_values),
        State=("State", collapse_values),
        Sources=("Source File", collapse_values),
        Source_Tabs=("Source Tab", collapse_values),
        Source_Types=("Source Type", collapse_values),
        Source_Count=("Source File", "nunique"),
        Source_Tab_Count=("Source Tab", "nunique"),
        Data_Text=("Data Text", collapse_values),
        Source_Text=("Search Text", collapse_values),
        Account_Name_Variants=("Account Name", collapse_values),
    ).reset_index()
    ranked = ranked.rename(
        columns={
            "Normalized Account Name": "Normalized Account Name",
            "Account_Name": "Account Name",
            "Parent_Health_System_if_available": "Parent Health System if available",
            "Source_Tabs": "Source Tabs",
            "Source_Types": "Source Types",
            "Source_Count": "Source Count",
            "Source_Tab_Count": "Source Tab Count",
            "Data_Text": "Data Text",
            "Source_Text": "Source Text",
            "Account_Name_Variants": "Account Name Variants",
        }
    )

    data_text = ranked["Data Text"].fillna("")
    source_text = ranked["Source Text"].fillna("")
    source_types = ranked["Source Types"].fillna("")
    known_epic_source = source_types.str.contains("EPIC Organization List", case=False, na=False)
    explicit_epic_signal = data_text.str.contains(r"\bepic\b", case=False, regex=True, na=False)

    ranked["New Epic System"] = data_text.str.contains(NEW_EPIC_PATTERN, na=False) | source_text.str.contains(NEW_EPIC_PATTERN, na=False)
    ranked["Existing Epic Customer"] = (known_epic_source | explicit_epic_signal) & ~ranked["New Epic System"]
    ranked["Imprivata Customer"] = source_text.str.contains(IMPRIVATA_PATTERN, na=False)
    ranked["Large Health System"] = source_text.str.contains(LARGE_HEALTH_SYSTEM_PATTERN, na=False)
    ranked["Large Physician Group"] = source_text.str.contains(LARGE_PHYSICIAN_GROUP_PATTERN, na=False)
    ranked["Non-Epic Migration Opportunity"] = (
        ~ranked["Existing Epic Customer"]
        & ~ranked["New Epic System"]
        & (
            source_text.str.contains(NON_EPIC_EHR_PATTERN, na=False)
            | source_text.str.contains(MIGRATION_PATTERN, na=False)
        )
    )
    ranked["Front-Door Workflow Fit Signal"] = source_text.str.contains(FRONT_DOOR_PATTERN, na=False)
    ranked["Security / Compliance Signal"] = source_text.str.contains(SECURITY_PATTERN, na=False)
    ranked["Buyer Role Signal"] = source_text.str.contains(BUYER_ROLE_PATTERN, na=False)
    ranked["Exclude Flag"] = source_text.str.contains(
        r"\b(?:exclude|do not contact|dnc|closed|inactive|duplicate only|competitor)\b",
        case=False,
        regex=True,
        na=False,
    )

    score_columns = ranked.apply(score_account, axis=1)
    ranked = pd.concat([ranked, score_columns], axis=1)
    ranked["Tier"] = ranked["AuthX Score"].map(assign_tier)
    ranked["Epic Status"] = ranked.apply(epic_status, axis=1)
    ranked["Imprivata Signal"] = ranked.apply(imprivata_signal, axis=1)
    ranked["Why This Account Matters"] = ranked.apply(why_account_matters, axis=1)
    ranked = pd.concat([ranked, ranked.apply(recommend_role, axis=1)], axis=1)
    ranked["Confidence"] = ranked.apply(confidence, axis=1)
    ranked["Missing Data"] = ranked.apply(missing_data, axis=1)

    tier_sort = {"Tier 1": 1, "Tier 2": 2, "Tier 3": 3, "Hold": 4, "Exclude": 5}
    ranked["Tier Sort"] = ranked["Tier"].map(tier_sort)
    ranked = ranked.sort_values(
        by=["Tier Sort", "AuthX Score", "New Epic System", "Existing Epic Customer", "Imprivata Customer", "Account Name"],
        ascending=[True, False, False, False, False, True],
    ).reset_index(drop=True)
    ranked.insert(0, "Rank", ranked.index + 1)

    output_columns = [
        "Rank",
        "Account Name",
        "Parent Health System if available",
        "Epic Status",
        "Imprivata Signal",
        "AuthX Score",
        "Tier",
        "Why This Account Matters",
        "Best Role to Pursue First",
        "Why This Role",
        "Secondary Roles",
        "Best Pitch",
        "Next Sales Action",
        "Confidence",
        "Missing Data",
        "Existing Epic Customer",
        "New Epic System",
        "Imprivata Customer",
        "Large Health System",
        "Large Physician Group",
        "Non-Epic Migration Opportunity",
        "Front-Door Workflow Fit Signal",
        "Security / Compliance Signal",
        "Buyer Role Signal",
        "Epic Status Points",
        "Timing Trigger Points",
        "Enterprise Value Points",
        "Imprivata / Competitive Signal Points",
        "Front-Door Workflow Fit Points",
        "Security / Compliance Urgency Points",
        "Buyer Role Fit Points",
        "EHR",
        "Status",
        "State",
        "Sources",
        "Source Tabs",
        "Source Types",
        "Account Name Variants",
    ]
    ranked = ranked[output_columns]

    summary = {
        "uploaded_workbooks": int(len(source_list)),
        "processed_workbooks": int(len(prepared)),
        "skipped_workbooks": int(len(skipped_sources)),
        "total_accounts": int(len(ranked)),
        "tier_1": int((ranked["Tier"] == "Tier 1").sum()),
        "tier_2": int((ranked["Tier"] == "Tier 2").sum()),
        "tier_3": int((ranked["Tier"] == "Tier 3").sum()),
        "hold": int((ranked["Tier"] == "Hold").sum()),
        "exclude": int((ranked["Tier"] == "Exclude").sum()),
        "existing_epic_customers": int(ranked["Existing Epic Customer"].sum()),
        "new_epic_systems": int(ranked["New Epic System"].sum()),
        "imprivata_customers": int(ranked["Imprivata Customer"].sum()),
        "large_health_systems": int(ranked["Large Health System"].sum()),
        "large_physician_groups": int(ranked["Large Physician Group"].sum()),
        "migration_opportunities": int(ranked["Non-Epic Migration Opportunity"].sum()),
    }
    return ranked, summary


def build_ranked_accounts(epic_file: BinaryIO, ehr_file: BinaryIO) -> tuple[pd.DataFrame, dict[str, int]]:
    sources = [
        SourceConfig("EPIC Organization list.xlsx", epic_file, "EPIC Organization List"),
        SourceConfig("Health Systems by EHR.xlsx", ehr_file, "Health Systems by EHR"),
    ]
    return build_ranked_accounts_from_sources(sources)


def scoring_summary_frame(summary: dict[str, int]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("Uploaded Workbooks", summary.get("uploaded_workbooks", 0)),
            ("Processed Workbooks", summary.get("processed_workbooks", 0)),
            ("Skipped Workbooks", summary.get("skipped_workbooks", 0)),
            ("Total Accounts", summary["total_accounts"]),
            ("Tier 1 Targets", summary["tier_1"]),
            ("Tier 2 Targets", summary["tier_2"]),
            ("Tier 3 Targets", summary["tier_3"]),
            ("Hold", summary["hold"]),
            ("Exclude", summary["exclude"]),
            ("Existing Epic Customers", summary["existing_epic_customers"]),
            ("New Epic Systems", summary["new_epic_systems"]),
            ("Imprivata Customers", summary["imprivata_customers"]),
            ("Large Health Systems", summary["large_health_systems"]),
            ("Large Physician Groups", summary["large_physician_groups"]),
            ("Non-Epic Migration Opportunities", summary["migration_opportunities"]),
        ],
        columns=["Metric", "Value"],
    )


def scoring_model_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("Epic Status", 25),
            ("Timing Trigger", 20),
            ("Enterprise Value", 15),
            ("Imprivata / Competitive Signal", 15),
            ("Front-Door Workflow Fit", 10),
            ("Security / Compliance Urgency", 10),
            ("Buyer Role Fit", 5),
        ],
        columns=["Score Component", "Max Points"],
    )


def dataframe_to_excel_bytes(ranked: pd.DataFrame, summary: dict[str, int]) -> bytes:
    output = BytesIO()
    role_strategy_columns = [
        "Account Name",
        "Tier",
        "AuthX Score",
        "Best Role to Pursue First",
        "Why This Role",
        "Secondary Roles",
        "Best Pitch",
        "Next Sales Action",
        "Confidence",
        "Missing Data",
    ]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        ranked.to_excel(writer, index=False, sheet_name="All Scored Accounts")
        ranked[ranked["Tier"] == "Tier 1"].to_excel(writer, index=False, sheet_name="Tier 1 Targets")
        ranked[ranked["Tier"] == "Tier 2"].to_excel(writer, index=False, sheet_name="Tier 2 Targets")
        ranked[role_strategy_columns].to_excel(writer, index=False, sheet_name="Role Strategy")

        summary_frame = scoring_summary_frame(summary)
        model_frame = scoring_model_frame()
        summary_frame.to_excel(writer, index=False, sheet_name="Scoring Summary", startrow=0)
        model_frame.to_excel(writer, index=False, sheet_name="Scoring Summary", startrow=len(summary_frame) + 3)

        workbook = writer.book
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            for row in sheet.iter_rows():
                if row[0].row in {1, len(summary_frame) + 4 if sheet.title == "Scoring Summary" else -1}:
                    for cell in row:
                        if cell.value is not None:
                            cell.fill = header_fill
                            cell.font = header_font
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                width = min(max(max_length + 2, 12), 55)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
            sheet.auto_filter.ref = sheet.dimensions

    output.seek(0)
    return output.read()
